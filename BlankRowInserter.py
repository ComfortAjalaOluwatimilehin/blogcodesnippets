# USAGE python blankRowInserter.py <N> <M> <excel sheet path>

import openpyxl
import os
import sys
import logging

logging.basicConfig(level=logging.DEBUG,
                    format="%(asctime)s – %(levelname)s – %(message)s")

if len(sys.argv) == 4:
    # TODO: extract N M and file path
    n, m, filepath = int(sys.argv[1]), int(sys.argv[2]), sys.argv[3]
    # TODO: read file
    try:
        wb = openpyxl.load_workbook(os.path.abspath(filepath))
        sheet = wb.get_active_sheet()
        # TODO loop M times starting at row N
        for row in range(n, n+m):
            for col in range(1, sheet.max_column+1):
                sheet.cell(row=row, column=col).value = ""
                wb.save(os.path.join(os.path.dirname(
                    os.path.abspath(filepath)), "blankrow.xlsx"))
        wb.close()
    except Exception as err:
        logging.error("Error while opening File: " + str(err))
else:
    logging.error(
        "USAGE python blankRowInserter.py <N> <M> <excel sheet path>")
