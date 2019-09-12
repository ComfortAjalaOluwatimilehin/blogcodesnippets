# USAGE python text2sheet <Several text files seperated by space>

import logging
import sys
import os
import openpyxl

logging.basicConfig(level=logging.DEBUG,
                    format="%(asctime)s - %(levelname)s - %(message)s")

if len(sys.argv) > 1:
    # TODO: EXTRACT TEXT FILES
    files = sys.argv[1:]
    # TODO:CREATE NEW WORKBOOK
    wb = openpyxl.Workbook()
    # TODO: CREATE NEW SHEET / GET ACTIVE SHEET
    ws = wb.active
    if not ws:
        ws = wb.create_sheet("Sheet1")
        # TODO: LOOP OVER FILE NUMBER
        for col in range(1, len(files) + 1):
            file = files[col - 1]
            file = os.path.abspath(file)
            # TODO: ERROR HANDLE FILE READING
            if not os.path.exists(file):
                logging.warning("File: " + file + " - does not exist")
            else:
                file = open(file)
                content = file.readlines()
                for row in range(1, len(content)):
                    # TODO: UPDATE ROWS AND COLUMNS
                    ws.cell(row=row, column=col).value = content[row - 1]
                    logging.info("File: " + files[col - 1] + " - was added")
            file.close()
            wb.save("Text2Sheet.xlsx")
            wb.close()
            logging.info("Workbook closed")
    else:
        logging.error(
            "USAGE python text2sheet <Several text files seperated by space>")
